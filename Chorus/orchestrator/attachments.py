from __future__ import annotations

import hashlib
import json
import os
import shutil
import subprocess
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from fastapi import HTTPException, UploadFile
from PIL import Image
from pydantic import BaseModel

from orchestrator.models import AttachmentRecord

TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".json", ".jsonl", ".csv", ".tsv", ".yaml", ".yml",
    ".xml", ".html", ".htm", ".py", ".ts", ".tsx", ".js", ".jsx", ".sql", ".log",
    ".ini", ".cfg", ".toml", ".sh", ".ps1", ".rb", ".go", ".java", ".rs", ".c", ".cpp",
    ".h", ".hpp",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}
AUDIO_EXTENSIONS = {".mp3", ".wav", ".m4a", ".aac", ".flac", ".ogg"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v"}

_WHISPER_MODEL_CACHE: dict[str, Any] = {}


class PreparedAttachment(BaseModel):
    attachment_id: str
    filename: str
    media_type: str
    size_bytes: int
    storage_path: str
    kind: str
    preview_text: str
    extracted_text: str
    metadata: dict[str, Any]


def _attachment_root() -> Path:
    raw = os.getenv("ORC_ATTACHMENT_DIR", "").strip()
    if raw:
        return Path(raw)
    return Path.cwd() / "uploads"


def _safe_filename(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in {".", "-", "_"} else "_" for ch in name.strip())
    return cleaned[:120] or "attachment.bin"


def _max_upload_bytes() -> int:
    return int(os.getenv("ORC_ATTACHMENT_MAX_BYTES", str(250 * 1024 * 1024)))


def _max_inline_chars() -> int:
    return int(os.getenv("ORC_ATTACHMENT_INLINE_CHARS", "20000"))


def _preview_chars() -> int:
    return int(os.getenv("ORC_ATTACHMENT_PREVIEW_CHARS", "800"))


def _truncate(text: str, limit: int) -> str:
    stripped = " ".join(text.split())
    if len(stripped) <= limit:
        return stripped
    return stripped[: limit - 1].rstrip() + "…"


def _kind_for_suffix(filename: str, media_type: str) -> str:
    suffix = Path(filename).suffix.lower()
    media = (media_type or "").lower()
    if suffix == ".pdf" or media == "application/pdf":
        return "pdf"
    if suffix == ".docx":
        return "docx"
    if suffix == ".pptx":
        return "pptx"
    if suffix == ".xlsx":
        return "xlsx"
    if suffix in TEXT_EXTENSIONS or media.startswith("text/"):
        return "text"
    if suffix in IMAGE_EXTENSIONS or media.startswith("image/"):
        return "image"
    if suffix in AUDIO_EXTENSIONS or media.startswith("audio/"):
        return "audio"
    if suffix in VIDEO_EXTENSIONS or media.startswith("video/"):
        return "video"
    return "binary"


async def persist_upload(upload: UploadFile, workspace_id: str) -> PreparedAttachment:
    filename = _safe_filename(upload.filename or "attachment.bin")
    media_type = (upload.content_type or "application/octet-stream").strip() or "application/octet-stream"
    root = _attachment_root()
    root.mkdir(parents=True, exist_ok=True)
    workspace_dir = root / workspace_id
    workspace_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.NamedTemporaryFile(delete=False, dir=workspace_dir, suffix=Path(filename).suffix) as tmp:
        temp_path = Path(tmp.name)
        digest = hashlib.sha256()
        size_bytes = 0
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            size_bytes += len(chunk)
            if size_bytes > _max_upload_bytes():
                tmp.close()
                temp_path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail="attachment_too_large")
            digest.update(chunk)
            tmp.write(chunk)

    attachment_id = digest.hexdigest()[:24]
    final_dir = workspace_dir / attachment_id
    final_dir.mkdir(parents=True, exist_ok=True)
    final_path = final_dir / filename
    if final_path.exists():
        temp_path.unlink(missing_ok=True)
    else:
        shutil.move(str(temp_path), str(final_path))

    kind = _kind_for_suffix(filename, media_type)
    extracted_text, metadata = extract_attachment_text(final_path, kind, media_type)
    preview_text = _truncate(extracted_text, _preview_chars())
    return PreparedAttachment(
        attachment_id=attachment_id,
        filename=filename,
        media_type=media_type,
        size_bytes=size_bytes,
        storage_path=str(final_path),
        kind=kind,
        preview_text=preview_text,
        extracted_text=_truncate(extracted_text, _max_inline_chars()),
        metadata=metadata,
    )


def build_attachment_record(prepared: PreparedAttachment, workspace_id: str) -> AttachmentRecord:
    return AttachmentRecord(
        attachment_id=prepared.attachment_id,
        workspace_id=workspace_id,
        filename=prepared.filename,
        media_type=prepared.media_type,
        kind=prepared.kind,
        size_bytes=prepared.size_bytes,
        storage_path=prepared.storage_path,
        preview_text=prepared.preview_text,
        extracted_text=prepared.extracted_text,
        metadata=prepared.metadata,
        created_at=time.time(),
    )


def build_attachment_context(attachments: list[AttachmentRecord]) -> str:
    blocks: list[str] = []
    for attachment in attachments:
        header = (
            f"[{attachment.attachment_id}] {attachment.filename} "
            f"({attachment.kind}, {attachment.media_type}, {attachment.size_bytes} bytes)"
        )
        body = attachment.extracted_text.strip() or attachment.preview_text.strip()
        if not body:
            body = "No text could be extracted automatically. Use the file metadata and user prompt."
        blocks.append(f"{header}\n{body}")
    return "\n\n".join(blocks)


def extract_attachment_text(path: Path, kind: str, media_type: str) -> tuple[str, dict[str, Any]]:
    metadata: dict[str, Any] = {"path": str(path), "media_type": media_type}
    try:
        if kind == "pdf":
            return _extract_pdf(path, metadata), metadata
        if kind == "docx":
            return _extract_docx(path, metadata), metadata
        if kind == "pptx":
            return _extract_pptx(path, metadata), metadata
        if kind == "xlsx":
            return _extract_xlsx(path, metadata), metadata
        if kind == "text":
            return _extract_text(path, metadata), metadata
        if kind == "image":
            return _extract_image(path, metadata), metadata
        if kind in {"audio", "video"}:
            return _extract_media(path, kind, metadata), metadata
    except Exception as exc:  # noqa: BLE001
        metadata["extraction_error"] = str(exc)
    return (
        f"Stored binary attachment {path.name}. Automatic text extraction is not available for this file type.",
        metadata,
    )


def _extract_pdf(path: Path, metadata: dict[str, Any]) -> str:
    try:
        import fitz  # type: ignore

        doc = fitz.open(path)
        metadata["pages"] = len(doc)
        parts: list[str] = []
        for index, page in enumerate(doc, start=1):
            text = page.get_text("text").strip()
            if text:
                parts.append(f"Page {index}\n{text}")
        if parts:
            return "\n\n".join(parts)
    except Exception:
        pass

    import pypdf

    reader = pypdf.PdfReader(str(path))
    metadata["pages"] = len(reader.pages)
    parts = []
    for index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            parts.append(f"Page {index}\n{text}")
    return "\n\n".join(parts) if parts else f"PDF stored as {path.name}, but no extractable text was found."


def _extract_docx(path: Path, metadata: dict[str, Any]) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            parts.append(line)
    metadata["paragraphs"] = len(parts)
    return "\n".join(parts) if parts else f"DOCX stored as {path.name}, but no text was extracted."


def _extract_pptx(path: Path, metadata: dict[str, Any]) -> str:
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    slides: list[str] = []
    with zipfile.ZipFile(path) as archive:
        slide_files = sorted(
            name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        metadata["slides"] = len(slide_files)
        for index, slide_name in enumerate(slide_files, start=1):
            xml = archive.read(slide_name)
            root = ET.fromstring(xml)
            texts = [node.text or "" for node in root.findall(".//a:t", ns)]
            slide_text = "\n".join(part.strip() for part in texts if part and part.strip())
            if slide_text:
                slides.append(f"Slide {index}\n{slide_text}")
    return "\n\n".join(slides) if slides else f"PPTX stored as {path.name}, but no text was extracted."


def _extract_xlsx(path: Path, metadata: dict[str, Any]) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    metadata["sheets"] = wb.sheetnames
    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                rows.append(" | ".join(values))
            if len(rows) >= 50:
                break
        if rows:
            sections.append(f"Sheet {sheet_name}\n" + "\n".join(rows))
    return "\n\n".join(sections) if sections else f"Workbook stored as {path.name}, but no text rows were extracted."


def _extract_text(path: Path, metadata: dict[str, Any]) -> str:
    raw = path.read_bytes()
    metadata["encoding"] = "utf-8"
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        metadata["encoding"] = "latin-1"
        return raw.decode("latin-1", errors="ignore")


def _extract_image(path: Path, metadata: dict[str, Any]) -> str:
    with Image.open(path) as image:
        metadata["width"] = image.width
        metadata["height"] = image.height
        metadata["mode"] = image.mode
    return (
        f"Image attachment {path.name} ({metadata['width']}x{metadata['height']}). "
        "Automatic OCR is not configured, so only image metadata is available."
    )


def _extract_media(path: Path, kind: str, metadata: dict[str, Any]) -> str:
    metadata["probe"] = _ffprobe(path)
    if os.getenv("ORC_ENABLE_MEDIA_TRANSCRIPTION", "1").strip().lower() in {"0", "false", "no", "off"}:
        return (
            f"{kind.capitalize()} attachment {path.name} stored. "
            "Automatic transcription is disabled; metadata only is available."
        )
    try:
        transcript = _transcribe_media(path)
        if transcript.strip():
            return transcript
    except Exception as exc:  # noqa: BLE001
        metadata["transcription_error"] = str(exc)
    return (
        f"{kind.capitalize()} attachment {path.name} stored. "
        "Metadata is available, but no transcript could be generated automatically."
    )


def _ffprobe(path: Path) -> dict[str, Any]:
    cmd = [
        "ffprobe",
        "-v",
        "error",
        "-print_format",
        "json",
        "-show_format",
        "-show_streams",
        str(path),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or "ffprobe_failed")
    return json.loads(proc.stdout or "{}")


def _transcribe_media(path: Path) -> str:
    from faster_whisper import WhisperModel

    model_name = os.getenv("ORC_TRANSCRIBE_MODEL", "base").strip() or "base"
    device = os.getenv("ORC_TRANSCRIBE_DEVICE", "auto").strip() or "auto"
    compute_type = os.getenv("ORC_TRANSCRIBE_COMPUTE_TYPE", "auto").strip() or "auto"
    model = _WHISPER_MODEL_CACHE.get(model_name)
    if model is None:
        model = WhisperModel(model_name, device=device, compute_type=compute_type)
        _WHISPER_MODEL_CACHE[model_name] = model

    with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as temp_audio:
        audio_path = Path(temp_audio.name)
    try:
        cmd = [
            "ffmpeg",
            "-y",
            "-i",
            str(path),
            "-ac",
            "1",
            "-ar",
            "16000",
            str(audio_path),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
        if proc.returncode != 0:
            raise RuntimeError(proc.stderr.strip() or "ffmpeg_extract_failed")
        segments, info = model.transcribe(str(audio_path))
        lines: list[str] = []
        for segment in segments:
            text = (segment.text or "").strip()
            if text:
                lines.append(text)
        prefix = f"Detected language: {getattr(info, 'language', 'unknown')}.\n"
        return prefix + " ".join(lines)
    finally:
        audio_path.unlink(missing_ok=True)
